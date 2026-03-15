import re
import sys
import hashlib
from datetime import datetime
from enum import Enum
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Categories ────────────────────────────────────────────────────────────────────

class Category(str, Enum):
    FOOD                   = 'Food'
    GROOMING               = 'Grooming'
    LONG_TRAVEL            = 'Long Travel'
    TRANSFER_EXTERNAL      = 'Transfer External'
    TRANSFER_INTERNAL      = 'Transfer Internal'
    HOME                   = 'Home'
    OTHER_ESSENTIALS       = 'Other Essentials'
    SUBSCRIPTION           = 'Subscription'
    LOCAL_TRAVEL           = 'Local Travel'
    REGULAR_PARTY_VACATION = 'Regular Party & Vacation'
    GYM                    = 'Gym'
    RENT                   = 'Rent'
    MEDICAL                = 'Medical'
    GROCERIES              = 'Groceries'
    TREATS_AND_GIFTS       = 'Treats & Gifts'
    OTHER_NON_ESSENTIALS   = 'Other Non Essentials'
    SPECIAL                = 'Special'
    BIG_PARTY_VACATION     = 'Big Party & Vacation'
    INVESTMENT             = 'Investment'
    PERSONAL               = 'Personal'


_CATEGORY_ALIASES: dict[str, Category] = {
    # investment
    'invest':        Category.INVESTMENT,
    'investment':    Category.INVESTMENT,
    'investments':   Category.INVESTMENT,
    # food / dining
    'food':          Category.FOOD,
    'dining':        Category.FOOD,
    'restaurant':    Category.FOOD,
    'zomato':        Category.FOOD,
    # groceries
    'grocery':       Category.GROCERIES,
    'groceries':     Category.GROCERIES,
    # travel
    'transport':     Category.LOCAL_TRAVEL,
    'travel':        Category.LOCAL_TRAVEL,
    'cab':           Category.LOCAL_TRAVEL,
    'uber':          Category.LOCAL_TRAVEL,
    'ola':           Category.LOCAL_TRAVEL,
    'airport':       Category.LONG_TRAVEL,
    'flight':        Category.LONG_TRAVEL,
    # medical
    'health':        Category.MEDICAL,
    'medical':       Category.MEDICAL,
    'medicine':      Category.MEDICAL,
    # rent
    'rent':          Category.RENT,
    # subscription
    'subscription':  Category.SUBSCRIPTION,
    # gym
    'gym':           Category.GYM,
}


def normalise_category(raw: str) -> str:
    """Map bracket text to a canonical Category value, or '' if unrecognised."""
    key = raw.strip().lower()
    # 1. Alias match
    cat = _CATEGORY_ALIASES.get(key)
    if cat:
        return cat.value
    # 2. Direct enum value match (case-insensitive)
    for member in Category:
        if member.value.lower() == key:
            return member.value
    # 3. Unknown — return '' so keyword scan can take over
    return ''


# ── Regex Patterns ────────────────────────────────────────────────────────────────

MSG_RE = re.compile(
    r'^\[(\d{2}/\d{2}/\d{2}),\s+(\d{1,2}:\d{2}:\d{2}[\s\u202f]+[AP]M)\]\s+([^:]+):\s+(.+)$',
    re.DOTALL
)

FOREIGN_SUFFIX_RE = re.compile(
    r'^\s*[+\-]?\s*(?:₹\s*)?\d[\d,]*(?:\.\d+)?\s*(?:usd|eur|gbp|sgd|aed|myr|thb)\b',
    re.IGNORECASE
)

EXPENSE_RE = re.compile(
    r'^([+\-])?\s*(?:₹\s*)?'
    r'(\d[\d,]*(?:\.\d+)?(?:\s*(?:k|lakh|lac))?)'
    r'\s+(.+)$',
    re.IGNORECASE
)

CATEGORY_RE  = re.compile(r'\[([^\]]+)\]')
EDITED_RE    = re.compile(r'[\s\u200e]*<This message was edited>\s*$')
DELETED_RE   = re.compile(r'(You deleted this message|This message was deleted)', re.IGNORECASE)

IGNORE_PATTERNS = [
    re.compile(p, re.IGNORECASE) for p in [
        r'^[\u200e\u202f]*Messages and calls are end-to-end encrypted',
        r'^[\u200e\u202f]*You (created|deleted|added|removed|changed|left)',
        r'^[\u200e\u202f]*(image|sticker|video|audio|document|GIF) omitted',
        r'^[\u200e\u202f]*This message was deleted',
        r'^[\u200e\u202f]*You deleted this message',
        r'^[\u200e\u202f]+$',
        r'^\s*$',
    ]
]

# ── Column layout ─────────────────────────────────────────────────────────────────
# Expenses sheet  : Date | Person | Amount | Category | Description | _key (hidden)
# Needs Review    : Line# | Date | Person | Raw Text | Reason | _key (hidden)
EXP_KEY_COL = 7   # column G
REV_KEY_COL = 6   # column F

# ── Helpers ───────────────────────────────────────────────────────────────────────

def make_key(*parts) -> str:
    raw = '|'.join(str(p).strip().lower() for p in parts)
    return hashlib.md5(raw.encode()).hexdigest()


def fmt_date(date_str: str) -> str:
    try:
        return datetime.strptime(date_str, '%d/%m/%y').strftime('%d %b %Y')
    except ValueError:
        return date_str


def normalise_amount(raw: str, sign) :
    raw = raw.strip().lower().replace(',', '')
    multiplier = 1.0
    if raw.endswith('lakh') or raw.endswith('lac'):
        raw = re.sub(r'(lakh|lac)$', '', raw).strip()
        multiplier = 100_000.0
    elif raw.endswith('k'):
        raw = raw[:-1].strip()
        multiplier = 1_000.0
    try:
        amount = float(raw) * multiplier
        return round(-amount if sign == '-' else amount, 2)
    except ValueError:
        return None


def extract_category(text: str):
    m = CATEGORY_RE.search(text)
    if m:
        return normalise_category(m.group(1).strip()), CATEGORY_RE.sub('', text).strip()
    return '', text


def is_ignored(body: str) -> bool:
    return any(p.search(body) for p in IGNORE_PATTERNS)

# ── Message grouping ──────────────────────────────────────────────────────────────

def group_messages(lines) :
    """Return list of (start_line_no, full_message_text)."""
    messages, current_no, current = [], None, None
    for i, raw in enumerate(lines, 1):
        stripped = raw.rstrip('\n').strip()
        if MSG_RE.match(stripped):
            if current is not None:
                messages.append((current_no, current))
            current_no, current = i, stripped
        elif current is not None and stripped:
            current += '\n' + stripped
    if current is not None:
        messages.append((current_no, current))
    return messages

# ── Core parser ───────────────────────────────────────────────────────────────────

def parse_chat(filepath: str) :
    parsed, unparsed = [], []

    with open(filepath, encoding='utf-8') as f:
        raw_lines = f.readlines()

    for line_no, msg in group_messages(raw_lines):
        msg = EDITED_RE.sub('', msg)
        m   = MSG_RE.match(msg)
        if not m:
            continue

        date_str, _, person, body = m.groups()
        person, body = person.strip(), body.strip()
        date_fmt = fmt_date(date_str)

        if is_ignored(body):
            continue

        for sub in body.splitlines():
            sub = sub.strip()
            if not sub or is_ignored(sub):
                continue
            if DELETED_RE.search(sub):
                continue

            # Foreign currency
            if FOREIGN_SUFFIX_RE.match(sub):
                unparsed.append({
                    'line_no': line_no, 'date': date_fmt, 'person': person,
                    'text': sub, 'reason': 'Foreign currency — manual review needed',
                    'key': make_key(date_fmt, person, sub),
                })
                continue

            # Must start with digit
            bare = sub.lstrip('+-').lstrip()
            if not bare or not bare[0].isdigit():
                unparsed.append({
                    'line_no': line_no, 'date': date_fmt, 'person': person,
                    'text': sub, 'reason': 'Does not start with a number',
                    'key': make_key(date_fmt, person, sub),
                })
                continue

            exp_m = EXPENSE_RE.match(sub)
            if exp_m:
                sign, amount_raw, text_raw = exp_m.groups()
                if sign == '-':
                    unparsed.append({
                        'line_no': line_no, 'date': date_fmt, 'person': person,
                        'text': sub, 'reason': 'Negative amount — manual review needed',
                        'key': make_key(date_fmt, person, sub),
                    })
                    continue
                amount = normalise_amount(amount_raw, sign)
                if amount is not None:
                    category, description = extract_category(text_raw)
                    if not category:
                        # Infer category from keywords in description
                        for word in re.findall(r'\b\w+\b', description.lower()):
                            if word in _CATEGORY_ALIASES:
                                category = _CATEGORY_ALIASES[word].value
                                break
                    if not category and re.search(r'\binvest(ment)?\b', description, re.IGNORECASE):
                        category = Category.INVESTMENT.value
                    parsed.append({
                        'line_no': line_no, 'date': date_fmt, 'person': person,
                        'amount': amount, 'is_credit': sign == '+',
                        'category': category, 'description': description,
                        'key': make_key(date_fmt, person, amount, description),
                    })
                else:
                    unparsed.append({
                        'line_no': line_no, 'date': date_fmt, 'person': person,
                        'text': sub, 'reason': 'Could not parse amount',
                        'key': make_key(date_fmt, person, sub),
                    })
            else:
                unparsed.append({
                    'line_no': line_no, 'date': date_fmt, 'person': person,
                    'text': sub, 'reason': 'No expense pattern matched',
                    'key': make_key(date_fmt, person, sub),
                })

    _stamp_occurrence_keys(parsed)
    _stamp_occurrence_keys(unparsed)
    return parsed, unparsed


def _stamp_occurrence_keys(rows):
    """
    Append an occurrence index to each key so that two identical expenses on
    the same day by the same person get distinct keys (_0, _1, _2 …).

    Stable across re-exports because WhatsApp always exports in chronological
    order — the Nth identical message always receives index N.
    """
    counts = {}
    for row in rows:
        base      = row['key']
        idx       = counts.get(base, 0)
        row['key'] = f'{base}_{idx}'
        counts[base] = idx + 1


# ── Read existing keys from xlsx ──────────────────────────────────────────────────

def load_existing_keys(xlsx_path: str) :
    """Return (expense_keys, review_keys) already present in the file."""
    expense_keys, review_keys = set(), set()
    try:
        wb = load_workbook(xlsx_path, read_only=True)
    except FileNotFoundError:
        return expense_keys, review_keys

    if 'Expenses' in wb.sheetnames:
        ws = wb['Expenses']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= EXP_KEY_COL and row[EXP_KEY_COL - 1]:
                expense_keys.add(str(row[EXP_KEY_COL - 1]).strip())

    if 'Needs Review' in wb.sheetnames:
        ws2 = wb['Needs Review']
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= REV_KEY_COL and row[REV_KEY_COL - 1]:
                review_keys.add(str(row[REV_KEY_COL - 1]).strip())

    wb.close()
    return expense_keys, review_keys

# ── Styling ───────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill('solid', start_color='1F3864')
ALT_FILL     = PatternFill('solid', start_color='EEF2FF')
ERR_FILL     = PatternFill('solid', start_color='FFF3CD')
ERR_HDR_FILL = PatternFill('solid', start_color='7B3F00')
thin         = Side(style='thin', color='D0D0D0')
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)


def _base_style(cell, fill, fmt, align):
    cell.font      = Font(name='Arial', size=10)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = BORDER
    if fmt:
        cell.number_format = fmt


def style_header(cell, fill=None):
    cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.fill      = fill or HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = BORDER


def style_data(cell, row_idx, fmt=None, align='left'):
    _base_style(cell, ALT_FILL if row_idx % 2 == 0 else PatternFill(), fmt, align)


def style_review(cell, fmt=None, align='left'):
    _base_style(cell, ERR_FILL, fmt, align)

# ── Sheet bootstrap (headers + freeze) ────────────────────────────────────────────

def _setup_expense_sheet(ws):
    ws.freeze_panes = 'A2'
    headers = ['Date', 'Person', 'Credit (₹)', 'Debit (₹)', 'Category', 'Description', '_key']
    widths  = [14,     18,       14,            14,           18,         52,             0.1  ]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        style_header(ws.cell(row=1, column=col, value=h))
        ws.column_dimensions[get_column_letter(col)].width = w
    # Hide key column
    ws.column_dimensions[get_column_letter(EXP_KEY_COL)].hidden = True
    ws.row_dimensions[1].height = 24


def _setup_review_sheet(ws2):
    ws2.freeze_panes = 'A2'
    headers = ['Line #', 'Date', 'Person', 'Raw Text', 'Reason', '_key']
    widths  = [8,        14,     18,        60,          36,       0.1  ]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        style_header(ws2.cell(row=1, column=col, value=h), fill=ERR_HDR_FILL)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.column_dimensions[get_column_letter(REV_KEY_COL)].hidden = True
    ws2.row_dimensions[1].height = 24

# ── Write / update xlsx ───────────────────────────────────────────────────────────

def _write_expense_row(ws, row_num: int, row: dict):
    credit = row['amount'] if row.get('is_credit') else None
    debit  = row['amount'] if not row.get('is_credit') else None
    ws.cell(row=row_num, column=1, value=row['date'])
    ws.cell(row=row_num, column=2, value=row['person'])
    ws.cell(row=row_num, column=3, value=credit)
    ws.cell(row=row_num, column=4, value=debit)
    ws.cell(row=row_num, column=5, value=row['category'])
    ws.cell(row=row_num, column=6, value=row['description'])
    ws.cell(row=row_num, column=7, value=row['key'])
    style_data(ws.cell(row=row_num, column=1), row_num, '@')
    style_data(ws.cell(row=row_num, column=2), row_num)
    style_data(ws.cell(row=row_num, column=3), row_num, '#,##0.00', 'right')
    style_data(ws.cell(row=row_num, column=4), row_num, '#,##0.00', 'right')
    style_data(ws.cell(row=row_num, column=5), row_num)
    style_data(ws.cell(row=row_num, column=6), row_num)
    ws.cell(row=row_num, column=7).font = Font(name='Arial', size=8, color='AAAAAA')


def _write_review_row(ws2, row_num: int, row: dict):
    ws2.cell(row=row_num, column=1, value=row.get('line_no', ''))
    ws2.cell(row=row_num, column=2, value=row.get('date',    ''))
    ws2.cell(row=row_num, column=3, value=row.get('person',  ''))
    ws2.cell(row=row_num, column=4, value=row.get('text',    ''))
    ws2.cell(row=row_num, column=5, value=row.get('reason',  ''))
    ws2.cell(row=row_num, column=6, value=row.get('key',     ''))
    style_review(ws2.cell(row=row_num, column=1), align='center')
    style_review(ws2.cell(row=row_num, column=2))
    style_review(ws2.cell(row=row_num, column=3))
    style_review(ws2.cell(row=row_num, column=4))
    style_review(ws2.cell(row=row_num, column=5))
    ws2.cell(row=row_num, column=6).font = Font(name='Arial', size=8, color='AAAAAA')



def update_xlsx(parsed: list, unparsed: list, xlsx_path: str):
    """
    Create xlsx_path if it doesn't exist, or append only new rows if it does.
    Deduplication is done via MD5 keys stored in a hidden column.
    """
    existing_exp_keys, existing_rev_keys = load_existing_keys(xlsx_path)
    is_new_file = len(existing_exp_keys) == 0 and len(existing_rev_keys) == 0

    new_parsed   = [r for r in parsed   if r['key'] not in existing_exp_keys]
    new_unparsed = [r for r in unparsed if r['key'] not in existing_rev_keys]

    # ── Load or create workbook ───────────────────────────────────────────────────
    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.active.title = 'Expenses'

    # ── Expenses sheet ────────────────────────────────────────────────────────────
    if 'Expenses' not in wb.sheetnames:
        ws = wb.create_sheet('Expenses', 0)
        _setup_expense_sheet(ws)
    else:
        ws = wb['Expenses']
        if is_new_file:
            _setup_expense_sheet(ws)

    exp_start = ws.max_row + 1

    for i, row in enumerate(new_parsed):
        _write_expense_row(ws, exp_start + i, row)

    # ── Needs Review sheet ────────────────────────────────────────────────────────
    if new_unparsed:
        if 'Needs Review' not in wb.sheetnames:
            ws2 = wb.create_sheet('Needs Review')
            _setup_review_sheet(ws2)
        else:
            ws2 = wb['Needs Review']
            if is_new_file:
                _setup_review_sheet(ws2)

        rev_start = ws2.max_row + 1
        for i, row in enumerate(new_unparsed):
            _write_review_row(ws2, rev_start + i, row)

    wb.save(xlsx_path)

    # ── Summary ───────────────────────────────────────────────────────────────────
    skipped_exp = len(parsed)   - len(new_parsed)
    skipped_rev = len(unparsed) - len(new_unparsed)

    print(f'✅  Saved → {xlsx_path}')
    print(f'   New expense rows   : {len(new_parsed):>4}   (skipped {skipped_exp} duplicates)')
    print(f'   New review rows    : {len(new_unparsed):>4}   (skipped {skipped_rev} duplicates)')
    print(f'   Total in file      : {len(existing_exp_keys) + len(new_parsed):>4} expenses  |  '
          f'{len(existing_rev_keys) + len(new_unparsed)} needs review')


# ── Main ──────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    chat_file = sys.argv[1] if len(sys.argv) > 1 else '_chat.txt'
    out_file  = sys.argv[2] if len(sys.argv) > 2 else 'expenses.xlsx'

    parsed, unparsed = parse_chat(chat_file)
    update_xlsx(parsed, unparsed, out_file)
