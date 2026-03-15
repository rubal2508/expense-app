import sys
import hashlib
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling ───────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', start_color='1F3864')
ALT_FILL    = PatternFill('solid', start_color='EEF2FF')
TOTAL_FILL  = PatternFill('solid', start_color='D9E1F2')
thin        = Side(style='thin', color='D0D0D0')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

INVESTMENT_VALUE = 'Investment'   # must match Category.INVESTMENT.value in parse_expenses.py


def _base_style(cell, fill, fmt, align):
    cell.font      = Font(name='Arial', size=10)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = BORDER
    if fmt:
        cell.number_format = fmt


def _style_header(cell):
    cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = BORDER


def _style_total(cell, fmt=None, align='left'):
    cell.font      = Font(bold=True, name='Arial', size=10)
    cell.fill      = TOTAL_FILL
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = BORDER
    if fmt:
        cell.number_format = fmt


# ── Data loading ──────────────────────────────────────────────────────────────────

def load_expenses(xlsx_path: str):
    """
    Read the Expenses sheet and return monthly totals per person.
    Excludes credits (Credit column non-null) and Investment category rows.

    Expenses columns: Date(1) Person(2) Credit(3) Debit(4) Category(5) Description(6) _key(7)
    """
    wb = load_workbook(xlsx_path, read_only=True)
    if 'Expenses' not in wb.sheetnames:
        wb.close()
        raise ValueError(f"No 'Expenses' sheet found in {xlsx_path}")

    monthly = {}  # (sort_key, label) -> defaultdict(float) of person -> total

    for row in wb['Expenses'].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] == 'TOTAL':
            continue
        date_val, person, credit, debit, category = row[0], row[1], row[2], row[3], row[4]
        if credit is not None or debit is None:
            continue
        if (category or '') == INVESTMENT_VALUE:
            continue
        try:
            dt = datetime.strptime(str(date_val).strip(), '%d %b %Y')
        except ValueError:
            continue
        sort_key = dt.strftime('%Y%m')
        label    = dt.strftime('%b %Y')
        bucket   = monthly.setdefault((sort_key, label), defaultdict(float))
        bucket[str(person).strip()] += float(debit)

    wb.close()
    return monthly


# ── Analysis workbook builder ─────────────────────────────────────────────────────

def build_analysis(monthly: dict, out_path: str):
    """Write a new xlsx with a single Analysis sheet."""
    if not monthly:
        print('⚠️  No qualifying expense data found — nothing to write.')
        return

    sorted_months = sorted(monthly.keys())
    all_persons   = sorted({p for m in monthly.values() for p in m})
    n_persons     = len(all_persons)

    wb = Workbook()
    ws = wb.active
    ws.title        = 'Analysis'
    ws.freeze_panes = 'B2'

    # ── Headers ───────────────────────────────────────────────────────────────────
    headers = ['Month'] + all_persons + ['Total']
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=1, column=col, value=h))
        ws.column_dimensions[get_column_letter(col)].width = 14 if col == 1 else 16
    ws.row_dimensions[1].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────────────
    for r_idx, key in enumerate(sorted_months, 2):
        data = monthly[key]
        fill = ALT_FILL if r_idx % 2 == 0 else PatternFill()

        _base_style(ws.cell(row=r_idx, column=1, value=key[1]), fill, None, 'left')

        row_total = 0.0
        for c_idx, person in enumerate(all_persons, 2):
            val = data.get(person, None)
            row_total += val or 0.0
            _base_style(ws.cell(row=r_idx, column=c_idx, value=val), fill, '#,##0.00', 'right')

        tot_cell = ws.cell(row=r_idx, column=n_persons + 2, value=row_total)
        tot_cell.font          = Font(bold=True, name='Arial', size=10)
        tot_cell.fill          = fill
        tot_cell.number_format = '#,##0.00'
        tot_cell.alignment     = Alignment(horizontal='right', vertical='center')
        tot_cell.border        = BORDER

    # ── Grand total row ───────────────────────────────────────────────────────────
    grand_row   = len(sorted_months) + 2
    grand_total = 0.0

    _style_total(ws.cell(row=grand_row, column=1, value='Total'))

    for c_idx, person in enumerate(all_persons, 2):
        col_total = sum(monthly[m].get(person, 0.0) for m in sorted_months)
        grand_total += col_total
        _style_total(ws.cell(row=grand_row, column=c_idx, value=col_total), '#,##0.00', 'right')

    _style_total(ws.cell(row=grand_row, column=n_persons + 2, value=grand_total), '#,##0.00', 'right')

    wb.save(out_path)
    print(f'✅  Analysis saved → {out_path}')
    print(f'   Months  : {len(sorted_months)}')
    print(f'   Persons : {", ".join(all_persons)}')


# ── Main ──────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    in_file  = sys.argv[1] if len(sys.argv) > 1 else 'expenses.xlsx'
    out_file = sys.argv[2] if len(sys.argv) > 2 else 'analysis.xlsx'

    monthly = load_expenses(in_file)
    build_analysis(monthly, out_file)
